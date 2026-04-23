import re
from datetime import datetime
from flask import Blueprint, request, jsonify

from modules.config import OUTPUT_EXCEL
from modules.scraper import _parse_filter_url, _fetch_issues_from_api, _format_issue_preview, gitlab_api_get, scrape_issue_api, resolve_project_id_from_url
from modules.excel_utils import issue_to_excel_row

excel_bp = Blueprint('excel_bp', __name__)

@excel_bp.route("/api/preview_issues", methods=["POST"])
def api_preview_issues():
    try:
        import requests as req_lib
    except ImportError:
        return jsonify({"error": "缺少 requests 套件，請執行：pip install requests"}), 500

    from modules.config import GITLAB_PRIVATE_TOKEN, GITLAB_PROJECT_ID

    body          = request.get_json(force=True)
    urls          = [u.strip() for u in (body.get("urls") or []) if u.strip()]
    
    client_project = body.get("project_id")
    project_id    = client_project if client_project else GITLAB_PROJECT_ID
    
    client_token  = (body.get("private_token") or "").strip()
    private_token = client_token if client_token else GITLAB_PRIVATE_TOKEN

    if not urls:
        return jsonify({"error": "請提供至少一個 Issue URL"}), 400

    headers = {"PRIVATE-TOKEN": private_token} if private_token else {}

    m_base = re.match(r"(https?://[^/]+)", urls[0])
    if not m_base:
        return jsonify({"error": "無法解析 base URL"}), 400
    base_url = m_base.group(1)

    # 嘗試從第一個網址解析 Project ID
    project_id = resolve_project_id_from_url(urls[0], private_token) or project_id

    if not project_id:
        return jsonify({"error": "請先在連線設定填寫 Project ID，或提供可解析的 Issue 網址"}), 400

    api_base = f"{base_url}/api/v4/projects/{project_id}/issues"
    issues   = []
    errors   = []

    for url in urls:
        m_iid = re.search(r"/(?:issues|work_items)/(\d+)(?:[/?#].*)?$", url)
        if not m_iid:
            errors.append({"url": url, "error": "無法解析 issue IID"})
            continue
        iid = int(m_iid.group(1))
        try:
            resp = gitlab_api_get(req_lib, f"{api_base}/{iid}", headers=headers, timeout=10)
            if resp.status_code == 401:
                return jsonify({"error": "401 未授權，請確認 API Token 有效"}), 401
            if resp.status_code == 404:
                errors.append({"url": url, "error": f"找不到 Issue #{iid}"})
                continue
            if not resp.ok:
                errors.append({"url": url, "error": f"API 錯誤 {resp.status_code}"})
                continue
            issues.append(_format_issue_preview(resp.json()))
        except Exception as e:
            errors.append({"url": url, "error": str(e)})

    return jsonify({
        "count":  len(issues),
        "issues": issues,
        "errors": errors,
    })


@excel_bp.route("/api/resolve_filter_url", methods=["POST"])
def api_resolve_filter_url():
    from modules.config import GITLAB_PRIVATE_TOKEN, GITLAB_PROJECT_ID
    
    body          = request.get_json()
    filter_url    = (body.get("filter_url")    or "").strip()
    
    client_project = body.get("project_id")
    project_id    = client_project if client_project else GITLAB_PROJECT_ID
    
    client_token  = (body.get("private_token") or "").strip()
    private_token = client_token if client_token else GITLAB_PRIVATE_TOKEN

    if not filter_url:
        return jsonify({"error": "請提供篩選頁面 URL"}), 400

    project_id = resolve_project_id_from_url(filter_url, private_token) or project_id

    try:
        base_url, project_path, api_params, labels, or_labels, not_labels = _parse_filter_url(filter_url)
        issues = _fetch_issues_from_api(
            base_url, project_path, api_params, labels, or_labels,
            not_labels=not_labels,
            project_id=project_id,
            private_token=private_token or None,
        )

        issue_list = [_format_issue_preview(issue) for issue in issues]

        return jsonify({
            "count":    len(issue_list),
            "project":  project_path,
            "issues":   issue_list,
            "web_urls": [i["web_url"] for i in issue_list],
        })

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        return jsonify({"error": f"未知錯誤：{e}"}), 500


@excel_bp.route("/api/batch_export_excel", methods=["POST"])
def api_batch_export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "缺少 openpyxl，請執行：pip install openpyxl"}), 500

    from modules.config import GITLAB_PRIVATE_TOKEN, GITLAB_PROJECT_ID

    body          = request.get_json()
    urls          = [u.strip() for u in (body.get("urls") or []) if u.strip()]
    
    client_project = body.get("project_id")
    project_id    = client_project if client_project else GITLAB_PROJECT_ID
    
    client_token  = (body.get("private_token") or "").strip()
    private_token = client_token if client_token else GITLAB_PRIVATE_TOKEN

    if not urls:
        return jsonify({"error": "請提供至少一個 Issue URL"}), 400
    project_id = resolve_project_id_from_url(urls[0], private_token) or project_id
    if not project_id:
        return jsonify({"error": "無法從網址解析且未提供 Project ID"}), 400

    m_base = re.match(r"(https?://[^/]+)", urls[0])
    if not m_base:
        return jsonify({"error": "無法解析 base URL"}), 400
    base_url = m_base.group(1)

    rows   = []
    errors = []

    for url in urls:
        m_iid = re.search(r"/(?:issues|work_items)/(\d+)(?:[/?#].*)?$", url)
        if not m_iid:
            errors.append({"url": url, "error": "無法解析 issue IID"})
            continue
        issue_iid = int(m_iid.group(1))
        try:
            issue = scrape_issue_api(base_url, int(project_id), issue_iid, private_token)
            rows.append(issue_to_excel_row(issue))
        except Exception as e:
            errors.append({"url": url, "error": str(e)})

    if not rows:
        return jsonify({"error": "所有 issue 均抓取失敗", "details": errors}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = "Issues"

    COLS = [
        "Issue ID", "Title", "State", "Priority", "Tag", "Epics", "其他標籤",
        "UI/UX", "FE", "BE", "Infra", "AI worker", "AI", "IE",
        "指派對象", "建立者", "建立時間", "最後更新",
        "Due Date", "預計完成版本", "Weight", "預估工時", "實花工時", "URL",
    ]

    col_widths = {
        "Issue ID": 10, "Title": 45, "State": 10, "Priority": 10,
        "Tag": 14, "Epics": 20, "其他標籤": 22,
        "UI/UX": 8, "FE": 8, "BE": 8, "Infra": 8, "AI worker": 10, "AI": 8, "IE": 8,
        "指派對象": 16, "建立者": 14, "建立時間": 16, "最後更新": 16,
        "Due Date": 13, "預計完成版本": 16, "Weight": 8,
        "預估工時": 10, "實花工時": 10, "URL": 60,
    }

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill  = PatternFill("solid", start_color="2F4F8F")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for ci, col_name in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font    = header_font
        cell.fill    = header_fill
        cell.alignment = header_align
        cell.border  = thin_border
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col_name, 14)

    ws.row_dimensions[1].height = 30

    data_font  = Font(name="Arial", size=10)
    fill_even  = PatternFill("solid", start_color="EEF2F8")
    fill_odd   = PatternFill("solid", start_color="FFFFFF")
    data_align = Alignment(vertical="center", wrap_text=False)
    url_font   = Font(name="Arial", size=10, color="1155CC", underline="single")

    priority_colors = {"High": "FF4444", "Medium": "FF9900", "Low": "44AA44"}

    for ri, row_data in enumerate(rows, start=2):
        fill = fill_even if ri % 2 == 0 else fill_odd
        for ci, col_name in enumerate(COLS, start=1):
            value = row_data.get(col_name, "")
            cell  = ws.cell(row=ri, column=ci, value=value)
            cell.border    = thin_border
            cell.alignment = data_align
            cell.fill      = fill

            if col_name == "URL" and value:
                cell.font = url_font
            elif col_name == "Priority" and value in priority_colors:
                cell.font = Font(name="Arial", size=10, bold=True,
                                 color=priority_colors[value])
            else:
                cell.font = data_font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"excel_{ts}_issues.xlsx"
    filepath = OUTPUT_EXCEL / filename
    wb.save(str(filepath))

    return jsonify({
        "filename": filename,
        "count":    len(rows),
        "errors":   errors,
    })
