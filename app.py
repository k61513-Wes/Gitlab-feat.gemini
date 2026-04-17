"""
app.py — Flask 後端（Selenium 爬蟲 + Gemini CLI subprocess）
新增：批次排程、自動存檔、統一輸出結構
版本：v1.2.0
"""

APP_VERSION = "v1.2.0"

import re
import time
import subprocess
import os
import shutil
import signal
import logging
from datetime import datetime
from pathlib import Path
from flask import Flask, request, jsonify, send_from_directory

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

app = Flask(__name__, static_folder=".")
logging.basicConfig(
    level=os.environ.get("APP_LOG_LEVEL", "INFO").upper(),
    format="%(asctime)s %(levelname)s [%(name)s] %(message)s",
)
logger = logging.getLogger("gitlab_issue_tool")

# ─── 常數設定 ────────────────────────────────────────────────────────────

GEMINI_CLI      = os.environ.get("GEMINI_CLI_PATH") or "gemini"
GEMINI_TIMEOUT  = int(os.environ.get("GEMINI_TIMEOUT", "300"))
GEMINI_PROBE_TIMEOUT = int(os.environ.get("GEMINI_PROBE_TIMEOUT", "12"))
MAX_INPUT_CHARS = int(os.environ.get("MAX_INPUT_CHARS", "40000"))
FLASK_HOST      = os.environ.get("FLASK_HOST", "127.0.0.1")
FLASK_PORT      = int(os.environ.get("FLASK_PORT", "5000"))
LLM_MODEL_PRIMARY    = os.environ.get("LLM_MODEL_PRIMARY", "gemini-2.5-pro").strip()
LLM_MODEL_FALLBACK_1 = os.environ.get("LLM_MODEL_FALLBACK_1", "gemma-4-31b-it").strip()
LLM_MODEL_FALLBACK_2 = os.environ.get("LLM_MODEL_FALLBACK_2", "gemma-4-26b-a4b-it").strip()

# 輸出資料夾（與 app.py 同層）
OUTPUT_DIR     = Path(os.environ.get("OUTPUT_DIR", "outputs"))
OUTPUT_RAW     = OUTPUT_DIR / "raw"
OUTPUT_RESULTS = OUTPUT_DIR / "results"
OUTPUT_EXCEL   = OUTPUT_DIR / "excel"
for _d in [OUTPUT_DIR, OUTPUT_RAW, OUTPUT_RESULTS, OUTPUT_EXCEL]:
    _d.mkdir(exist_ok=True)

# Prompt 模板資料夾
PROMPTS_DIR = Path("prompts")
PROMPTS_DIR.mkdir(exist_ok=True)

# 統一結構的固定區塊（需求 3）
REQUIRED_SECTIONS = ["問題說明", "根本原因", "解決方式", "待辦事項", "討論共識", "補充內容"]

# 強制結構的 system prompt（需求 3）
PROCESS_SYSTEM_PROMPT = """\
你是技術專案助理，請閱讀以下 GitLab Issue，整理成繁體中文的結構化摘要。

【輸出格式規定】
請嚴格依照以下六個區塊輸出，每個區塊都必須存在，不可省略：

## 問題說明
（描述 Issue 的背景與問題）

## 根本原因
（分析問題的根本原因，若不清楚請填「尚未確認」）

## 解決方式
（說明已採取或建議的解決方式，若尚無請填「尚未決定」）

## 待辦事項
（條列式列出待辦，若無請填「無」）

## 討論共識
（整理留言中的討論結論，若無留言請填「無」）

## 補充內容
（無法歸入以上分類的其他資訊，若無請填「無」）

注意：以上六個 ## 標題必須全部出現，順序不變，內容不足時填「無」或「尚未確認」。\
"""

MODEL_CHAIN_SPECS = [
    {"order": 1, "label": "Gemini 2.5 Pro", "model_id": LLM_MODEL_PRIMARY},
    {"order": 2, "label": "Gemma 4 31B", "model_id": LLM_MODEL_FALLBACK_1},
    {"order": 3, "label": "Gemma 4 26B", "model_id": LLM_MODEL_FALLBACK_2},
]


# ─── 存檔工具 ────────────────────────────────────────────────────────────

def _parse_gitlab_url_parts(url: str) -> dict:
    path = ""
    try:
        from urllib.parse import urlparse
        path = urlparse(url).path or ""
    except Exception:
        path = url or ""

    m = re.search(r"(?P<repo_path>.+?)/-/(?P<kind>issues|work_items)/(?P<number>\d+)$", path)
    if not m:
        return {
            "repo_name": "unknown-repo",
            "item_number": "unknown",
            "item_kind": "unknown",
        }

    repo_parts = [part for part in m.group("repo_path").split("/") if part]
    repo_name = repo_parts[-1] if repo_parts else "unknown-repo"
    if repo_name == "gitlab-profile" and len(repo_parts) >= 2:
        repo_name = repo_parts[-2]

    return {
        "repo_name": repo_name,
        "item_number": m.group("number"),
        "item_kind": m.group("kind"),
    }


def _sanitize_filename_part(value: str) -> str:
    value = (value or "").strip().replace("/", "-").replace("\\", "-")
    value = re.sub(r"[^\w\.-]+", "-", value, flags=re.UNICODE)
    value = re.sub(r"-{2,}", "-", value).strip("-._")
    return value or "unknown"


def build_output_filename(url: str, model_name: str = None, kind: str = "result", ext: str = "txt") -> str:
    parsed = _parse_gitlab_url_parts(url)
    repo_name = _sanitize_filename_part(parsed["repo_name"])
    item_number = _sanitize_filename_part(parsed["item_number"])
    date_str = datetime.now().strftime("%Y%m%d")
    if kind == "raw":
        model_part = "raw"
    else:
        model_part = _sanitize_filename_part(model_name or "unknown-model")
    return f"{repo_name}_{item_number}_{model_part}_{date_str}.{ext}"


def save_output(content: str, kind: str, url: str, model_name: str = None) -> str:
    """
    將內容寫入對應的子資料夾。
    kind: "raw"    → outputs/raw/
          "result" → outputs/results/
    回傳：寫入的檔案路徑（字串）
    """
    filename = build_output_filename(url, model_name=model_name, kind=kind, ext="txt")
    subdir   = OUTPUT_RAW if kind == "raw" else OUTPUT_RESULTS
    filepath = subdir / filename
    if filepath.exists():
        stem = filepath.stem
        suffix = filepath.suffix
        counter = 2
        while filepath.exists():
            filepath = subdir / f"{stem}_{counter}{suffix}"
            counter += 1
    filepath.write_text(content, encoding="utf-8")
    return str(filepath)


def gitlab_api_get(req_lib, url: str, **kwargs):
    """GitLab 內網 API 請求需忽略系統代理，避免被導到本機 proxy。"""
    session = req_lib.Session()
    session.trust_env = False
    try:
        return session.get(url, **kwargs)
    finally:
        session.close()


def list_outputs() -> list:
    """列出 raw、results、excel 下所有輸出檔，依時間倒序。"""
    all_files = (
        list(OUTPUT_RAW.glob("*.txt"))
        + list(OUTPUT_RESULTS.glob("*.txt"))
        + list(OUTPUT_EXCEL.glob("*.xlsx"))
    )
    all_files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    result = []
    for f in all_files:
        result.append({
            "filename": f.name,
            "size":     f.stat().st_size,
            "mtime":    datetime.fromtimestamp(f.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
            "kind":     "raw" if f.parent.name == "raw" else "excel" if f.parent.name == "excel" else "result",
        })
    return result


# ─── 結構驗證（需求 3） ──────────────────────────────────────────────────

def enforce_structure(text: str) -> str:
    """
    確保輸出包含六個必要區塊。
    若 Gemini 漏掉某個區塊，自動補上（填「無」）。
    """
    for section in REQUIRED_SECTIONS:
        pattern = rf"^##\s+{re.escape(section)}"
        if not re.search(pattern, text, re.MULTILINE):
            text = text.rstrip() + f"\n\n## {section}\n無\n"
    return text.strip()


# ─── Gemini CLI 呼叫 ─────────────────────────────────────────────────────

def _sanitize_text(text: str) -> str:
    return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", "", text)


def is_disallowed_model(model_name: str) -> bool:
    return "flash" in (model_name or "").strip().lower()


def get_model_chain() -> list:
    chain = []
    for item in MODEL_CHAIN_SPECS:
        model_id = (item.get("model_id") or "").strip()
        configured = bool(model_id)
        disallowed = is_disallowed_model(model_id)
        if not configured:
            reason = "unconfigured"
        elif disallowed:
            reason = "flash_not_allowed"
        else:
            reason = "ok"
        chain.append({
            "order": item["order"],
            "label": item["label"],
            "model_id": model_id,
            "configured": configured,
            "allowed": configured and not disallowed,
            "reason": reason,
        })
    return chain


def _resolve_gemini_executable() -> str:
    return shutil.which(GEMINI_CLI) or GEMINI_CLI


def _build_gemini_command(extra_args=None) -> list:
    extra_args = extra_args or []
    cli_path = _resolve_gemini_executable()
    if os.name == "nt" and str(cli_path).lower().endswith((".cmd", ".bat")):
        return ["cmd", "/c", cli_path, *extra_args]
    return [cli_path, *extra_args]


def _terminate_process_tree(process: subprocess.Popen) -> None:
    if process.poll() is not None:
        return
    if os.name == "nt":
        try:
            process.kill()
        except Exception:
            pass
        subprocess.run(
            ["taskkill", "/PID", str(process.pid), "/T", "/F"],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            timeout=2,
        )
    else:
        try:
            os.killpg(os.getpgid(process.pid), signal.SIGKILL)
        except ProcessLookupError:
            return


def _run_command_with_timeout(command: list, timeout: int, input_text: str = None) -> subprocess.CompletedProcess:
    popen_kwargs = {
        "stdin": subprocess.PIPE if input_text is not None else None,
        "stdout": subprocess.PIPE,
        "stderr": subprocess.PIPE,
        "text": True,
        "encoding": "utf-8",
    }
    if os.name == "nt":
        popen_kwargs["creationflags"] = subprocess.CREATE_NEW_PROCESS_GROUP
    else:
        popen_kwargs["start_new_session"] = True

    process = subprocess.Popen(command, **popen_kwargs)
    logger.info("run_command start timeout=%ss command=%s", timeout, command)
    try:
        stdout, stderr = process.communicate(input=input_text, timeout=timeout)
        logger.info("run_command done returncode=%s stdout_len=%s stderr_len=%s", process.returncode, len(stdout or ""), len(stderr or ""))
        return subprocess.CompletedProcess(
            args=command,
            returncode=process.returncode,
            stdout=stdout,
            stderr=stderr,
        )
    except subprocess.TimeoutExpired as exc:
        stdout = exc.stdout or ""
        stderr = exc.stderr or ""
        logger.warning("run_command timeout timeout=%ss pid=%s stdout_len=%s stderr_len=%s", timeout, getattr(process, "pid", None), len(stdout), len(stderr))
        try:
            _terminate_process_tree(process)
        except Exception:
            logger.exception("run_command timeout terminate_process_tree failed")
            pass
        raise subprocess.TimeoutExpired(
            cmd=command,
            timeout=timeout,
            output=stdout,
            stderr=stderr,
        )


def call_gemini_cli(system_prompt: str, user_text: str, timeout: int = None, model_name: str = None) -> str:
    system_prompt = _sanitize_text(system_prompt)
    user_text     = _sanitize_text(user_text)
    effective_timeout = timeout if (timeout and timeout > 0) else GEMINI_TIMEOUT
    selected_model = (model_name or "").strip()

    if selected_model and is_disallowed_model(selected_model):
        raise RuntimeError("不接受 Flash 模型，請改用 Gemini 2.5 Pro 或指定的 Gemma 4 模型")

    if len(user_text) > MAX_INPUT_CHARS:
        user_text = user_text[:MAX_INPUT_CHARS] + "\n\n[... 內容過長，已截斷 ...]"

    full_prompt = f"{system_prompt}\n\n---\n\n{user_text}"
    extra_args = ["--model", selected_model] if selected_model else []
    logger.info("llm_call start model=%s timeout=%ss prompt_len=%s user_text_len=%s", selected_model or "<default>", effective_timeout, len(system_prompt), len(user_text))

    try:
        result = _run_command_with_timeout(
            _build_gemini_command(extra_args),
            timeout=effective_timeout,
            input_text=full_prompt,
        )

        if result.returncode != 0:
            stderr = result.stderr.strip() or "（無 stderr 輸出）"
            logger.error("llm_call nonzero_exit model=%s returncode=%s stderr=%s", selected_model or "<default>", result.returncode, stderr[:500])
            raise RuntimeError(f"Gemini CLI 回傳錯誤碼 {result.returncode}：{stderr}")

        output = result.stdout.strip()
        if not output:
            logger.error("llm_call empty_output model=%s", selected_model or "<default>")
            raise RuntimeError("Gemini CLI 無輸出，請確認 CLI 已正確安裝並完成授權")

        logger.info("llm_call success model=%s output_len=%s", selected_model or "<default>", len(output))
        return output

    except subprocess.TimeoutExpired:
        logger.warning("llm_call timeout model=%s timeout=%ss", selected_model or "<default>", effective_timeout)
        raise RuntimeError(
            f"Gemini CLI 執行超時（>{effective_timeout}s），"
            "請縮短輸入長度或在連線設定調整逾時秒數"
        )
    except FileNotFoundError:
        logger.exception("llm_call cli_not_found cli=%s", GEMINI_CLI)
        raise RuntimeError(
            f"找不到 Gemini CLI 執行檔（{GEMINI_CLI!r}），"
            "請先安裝：npm install -g @google/gemini-cli，"
            "或設定 GEMINI_CLI_PATH 環境變數指向正確路徑"
        )


def probe_gemini_model(model_name: str, timeout: int = None) -> dict:
    effective_timeout = timeout if (timeout and timeout > 0) else GEMINI_PROBE_TIMEOUT
    prompt = "Please reply with OK only."
    logger.info("probe_model start model=%s timeout=%ss", model_name, effective_timeout)
    if is_disallowed_model(model_name):
        logger.warning("probe_model flash_not_allowed model=%s", model_name)
        return {
            "model": model_name,
            "ok": False,
            "status": "flash_not_allowed",
            "returncode": None,
            "stdout": "",
            "stderr": "Flash 模型不在允許清單內",
            "timeout": effective_timeout,
        }
    cmd = _build_gemini_command(["--model", model_name, "--prompt", prompt])

    try:
        result = _run_command_with_timeout(
            cmd,
            timeout=effective_timeout,
        )
        stdout = (result.stdout or "").strip()
        stderr = (result.stderr or "").strip()
        if result.returncode != 0:
            logger.warning("probe_model nonzero_exit model=%s returncode=%s stderr=%s", model_name, result.returncode, stderr[:500])
            return {
                "model": model_name,
                "ok": False,
                "status": "nonzero_exit",
                "returncode": result.returncode,
                "stdout": stdout[:200],
                "stderr": stderr[:500],
                "timeout": effective_timeout,
            }
        if not stdout:
            logger.warning("probe_model empty_output model=%s", model_name)
            return {
                "model": model_name,
                "ok": False,
                "status": "empty_output",
                "returncode": result.returncode,
                "stdout": "",
                "stderr": stderr[:500],
                "timeout": effective_timeout,
            }
        logger.info("probe_model success model=%s stdout=%s", model_name, stdout[:120])
        return {
            "model": model_name,
            "ok": True,
            "status": "ok",
            "returncode": result.returncode,
            "stdout": stdout[:200],
            "stderr": stderr[:200],
            "timeout": effective_timeout,
        }
    except subprocess.TimeoutExpired:
        logger.warning("probe_model timeout model=%s timeout=%ss", model_name, effective_timeout)
        return {
            "model": model_name,
            "ok": False,
            "status": "timeout",
            "returncode": None,
            "stdout": "",
            "stderr": f"probe timeout after {effective_timeout}s",
            "timeout": effective_timeout,
        }
    except FileNotFoundError:
        logger.exception("probe_model cli_not_found cli=%s model=%s", GEMINI_CLI, model_name)
        return {
            "model": model_name,
            "ok": False,
            "status": "cli_not_found",
            "returncode": None,
            "stdout": "",
            "stderr": f"CLI not found: {GEMINI_CLI}",
            "timeout": effective_timeout,
        }


# ─── Selenium 工具 ───────────────────────────────────────────────────────

def make_driver() -> webdriver.Chrome:
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1400,900")
    options.add_argument("--lang=zh-TW")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
    return webdriver.Chrome(options=options)


def selenium_login(driver, base_url, username, password):
    login_url = f"{base_url}/users/sign_in"
    driver.get(login_url)
    wait = WebDriverWait(driver, 15)

    user_input = wait.until(EC.presence_of_element_located((By.ID, "user_login")))
    user_input.clear()
    user_input.send_keys(username)

    pass_input = driver.find_element(By.ID, "user_password")
    pass_input.clear()
    pass_input.send_keys(password)
    pass_input.submit()

    try:
        wait.until(EC.url_changes(login_url))
    except TimeoutException:
        pass

    if "sign_in" in driver.current_url:
        raise ValueError("登入失敗，請確認帳號密碼")


def scroll_to_load_all(driver):
    last_height = driver.execute_script("return document.body.scrollHeight")
    for _ in range(25):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1.2)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(0.8)


def scrape_issue_selenium(base_url, username, password, url):
    driver = make_driver()
    try:
        selenium_login(driver, base_url, username, password)
        driver.get(url)
        wait = WebDriverWait(driver, 20)

        for selector in [
            "h1, .title, [data-testid='issue-title'], .work-item-title",
            ".work-item-description, .description, .detail-page-description",
            ".work-item-notes, .notes, .main-notes-list, #notes",
        ]:
            try:
                wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, selector)))
            except TimeoutException:
                pass

        scroll_to_load_all(driver)

        try:
            wait.until(lambda d: any(
                el.text.strip()
                for el in d.find_elements(By.CSS_SELECTOR, ".note-header-author-name")
            ))
        except TimeoutException:
            pass

        time.sleep(2.5)

        title = driver.execute_script(
            "var el = document.querySelector('h1');"
            "return el ? el.innerText.trim() : '';"
        ) or "（無標題）"

        description = driver.execute_script(
            "var el = document.querySelector('.work-item-description') "
            "|| document.querySelector('.description');"
            "return el ? el.innerText.trim() : '';"
        ) or ""

        labels = driver.execute_script(
            "var seen = {}; var out = [];"
            "document.querySelectorAll('.gl-label-text').forEach(function(el){"
            "  var t = el.innerText.trim();"
            "  if (t && !seen[t]) { seen[t]=true; out.push(t); }"
            "});"
            "return out;"
        ) or []

        js_code = """
var out = [];
function getAuthorAndTime(body) {
    var noteBody = body.parentElement;
    var grandParent = noteBody ? noteBody.parentElement : null;
    var author = 'unknown'; var created_at = '';
    if (grandParent) {
        var noteHeader = grandParent.querySelector('.note-header');
        if (noteHeader) {
            var spans = noteHeader.querySelectorAll('span');
            for (var s = 0; s < spans.length; s++) {
                if (spans[s].className.indexOf('note-header-author-name') >= 0) {
                    author = spans[s].innerText.trim(); break;
                }
            }
            var tEl = noteHeader.querySelector('time');
            if (tEl) {
                var dt = tEl.getAttribute('datetime');
                if (dt) {
                    var d = new Date(dt);
                    var tw = new Date(d.getTime() + 8*60*60*1000);
                    var pad = function(n){return n<10?'0'+n:''+n;};
                    created_at = tw.getUTCFullYear()+'-'+pad(tw.getUTCMonth()+1)+'-'+
                        pad(tw.getUTCDate())+' '+pad(tw.getUTCHours())+':'+pad(tw.getUTCMinutes());
                }
            }
        }
    }
    return {author:author, created_at:created_at};
}
function getBodyText(body) {
    var noteEl = body.querySelector('.note-text') || body;
    return noteEl.innerText.trim();
}
var entries = document.querySelectorAll('.timeline-entry.note-discussion');
for (var e = 0; e < entries.length; e++) {
    var entry = entries[e];
    var bodies = entry.querySelectorAll('.timeline-discussion-body');
    if (bodies.length === 0) continue;
    var mainInfo = getAuthorAndTime(bodies[0]);
    var mainText = getBodyText(bodies[0]);
    if (!mainText || (mainInfo.author==='unknown' && mainInfo.created_at==='')) continue;
    var replies = [];
    for (var r = 1; r < bodies.length; r++) {
        var replyInfo = getAuthorAndTime(bodies[r]);
        var replyText = getBodyText(bodies[r]);
        if (replyText && !(replyInfo.author==='unknown' && replyInfo.created_at==='')) {
            replies.push({author:replyInfo.author, created_at:replyInfo.created_at, body:replyText});
        }
    }
    out.push({author:mainInfo.author, created_at:mainInfo.created_at, body:mainText, replies:replies});
}
return out;
"""
        raw_comments = driver.execute_script(js_code) or []

        def clean_body(text):
            lines = [l for l in text.splitlines()
                     if not re.match(r"^/?cc\.?\s*@", l.strip())]
            return "\n".join(lines).strip()

        comments = []
        seen_bodies = set()
        for c in raw_comments:
            body_clean = clean_body(c.get("body", ""))
            if not body_clean or body_clean in seen_bodies:
                continue
            seen_bodies.add(body_clean)
            replies = [
                {"author": r.get("author", "unknown"),
                 "created_at": r.get("created_at", ""),
                 "body": clean_body(r.get("body", ""))}
                for r in c.get("replies", [])
                if clean_body(r.get("body", ""))
            ]
            comments.append({
                "author":     c.get("author", "unknown"),
                "created_at": c.get("created_at", ""),
                "body":       body_clean,
                "replies":    replies,
            })

        return {
            "url": url, "title": title,
            "author": "unknown", "created_at": "",
            "labels": labels, "description": description, "comments": comments,
        }
    finally:
        driver.quit()


def issue_to_text(issue: dict) -> str:
    iid   = issue.get("iid", "")
    title = issue.get("title", "（無標題）")
    lines = [f"# {'#' + str(iid) + ' ' if iid else ''}{title}", ""]

    # ── Metadata 區塊 ─────────────────────────────────────────────────────
    meta_lines = []
    if issue.get("state"):
        meta_lines.append(f"- **狀態**：{issue['state']}")
    if issue.get("author"):
        meta_lines.append(f"- **建立者**：{issue['author']}（{issue.get('created_at', '')}）")
    if issue.get("assignees"):
        meta_lines.append(f"- **指派對象**：{', '.join(issue['assignees'])}")
    if issue.get("labels"):
        meta_lines.append(f"- **標籤**：{', '.join(issue['labels'])}")
    ms = issue.get("milestone") or {}
    if ms.get("title"):
        ms_str = ms["title"]
        if ms.get("due_date"):
            ms_str += f"（截止 {ms['due_date']}）"
        meta_lines.append(f"- **Milestone**：{ms_str}")
    if issue.get("due_date"):
        meta_lines.append(f"- **Due Date**：{issue['due_date']}")
    if issue.get("weight") not in ("", None):
        meta_lines.append(f"- **Weight**：{issue['weight']}")
    if issue.get("time_estimate") or issue.get("time_spent"):
        meta_lines.append(
            f"- **時間追蹤**：預估 {issue.get('time_estimate') or '—'}｜"
            f"實花 {issue.get('time_spent') or '—'}"
        )
    if issue.get("updated_at"):
        meta_lines.append(f"- **最後更新**：{issue['updated_at']}")
    if meta_lines:
        lines += ["## 基本資訊"] + meta_lines + [""]

    # ── 說明內文 ──────────────────────────────────────────────────────────
    desc = issue.get("description") or ""
    if desc:
        keep_sections = ["功能描述", "建議方案", "補充內容"]
        parts = re.split(r"(?m)^(#+\s+.+)$", desc)
        current_title, current_body, sections = "", [], []
        for part in parts:
            if re.match(r"^#+\s+", part):
                if current_title:
                    sections.append((current_title, "\n".join(current_body).strip()))
                current_title, current_body = part.strip(), []
            else:
                current_body.append(part)
        if current_title:
            sections.append((current_title, "\n".join(current_body).strip()))

        found_any = False
        for title, body in sections:
            title_text = re.sub(r"^#+\s*", "", title)
            if any(kw in title_text for kw in keep_sections):
                lines += [f"## {title_text}", body, ""]
                found_any = True
        if not found_any:
            lines += ["## 說明", desc, ""]
    else:
        lines += ["## 說明", "（無說明）", ""]

    # ── 留言 ──────────────────────────────────────────────────────────────
    comments = issue.get("comments", [])
    if comments:
        lines.append(f"## 留言（共 {len(comments)} 則）")
        for i, c in enumerate(comments, 1):
            lines += [f"\n### 留言 {i}｜{c['author']}（{c['created_at']}）", c["body"]]
            for j, r in enumerate(c.get("replies", []), 1):
                lines += [
                    f"\n    ↳ 回覆 {j}｜{r['author']}（{r['created_at']}）",
                    "    " + r["body"].replace("\n", "\n    "),
                ]
    else:
        lines += ["## 留言", "（無留言）"]

    return "\n".join(lines)


# ─── GitLab API 版爬蟲（取代 Selenium，速度快 10 倍，只需 token）──────────

def scrape_issue_api(base_url: str, project_id: int, issue_iid: int,
                     private_token: str = None) -> dict:
    """
    使用 GitLab REST API 取得 Issue 詳情與留言。
    回傳格式與 scrape_issue_selenium 完全相同，可直接接入 issue_to_text()。
    """
    try:
        import requests as req_lib
    except ImportError:
        raise RuntimeError("缺少 requests 套件，請執行：pip install requests")

    headers = {}
    if private_token:
        headers["PRIVATE-TOKEN"] = private_token

    api_base = f"{base_url}/api/v4/projects/{project_id}"

    # ── 取 Issue 主體 ──────────────────────────────────────────────────────
    issue_resp = gitlab_api_get(
        req_lib,
        f"{api_base}/issues/{issue_iid}",
        headers=headers,
        timeout=15,
    )
    if issue_resp.status_code == 401:
        raise RuntimeError("401 未授權，請確認 API Token 有效")
    if issue_resp.status_code == 404:
        raise RuntimeError(
            f"找不到 Issue #{issue_iid}（project_id={project_id}），"
            "請確認 Project ID 正確"
        )
    if not issue_resp.ok:
        raise RuntimeError(
            f"API 錯誤 {issue_resp.status_code}：{issue_resp.text[:200]}"
        )
    issue_data = issue_resp.json()

    # ── 取留言（notes），自動翻頁，過濾 system notes ───────────────────────
    all_notes = []
    page = 1
    while True:
        notes_resp = gitlab_api_get(
            req_lib,
            f"{api_base}/issues/{issue_iid}/notes",
            params={"per_page": 100, "sort": "asc", "page": page},
            headers=headers, timeout=15,
        )
        if not notes_resp.ok:
            break
        batch = notes_resp.json()
        if not batch:
            break
        all_notes.extend(batch)
        if len(batch) < 100:
            break
        page += 1

    # ── 時間轉換（UTC → UTC+8）──────────────────────────────────────────────
    def fmt_time(ts: str) -> str:
        if not ts:
            return ""
        try:
            from datetime import datetime, timezone, timedelta
            dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
            tw = dt.astimezone(timezone(timedelta(hours=8)))
            return tw.strftime("%Y-%m-%d %H:%M")
        except Exception:
            return ts[:16].replace("T", " ")

    # ── 格式化留言 ─────────────────────────────────────────────────────────
    comments = []
    seen_bodies: set = set()
    for note in all_notes:
        if note.get("system", False):
            continue  # 跳過 label/assignee 等系統訊息
        body = note.get("body", "").strip()
        if not body or body in seen_bodies:
            continue
        seen_bodies.add(body)
        comments.append({
            "author":     note.get("author", {}).get("name", "unknown"),
            "created_at": fmt_time(note.get("created_at", "")),
            "body":       body,
            "replies":    [],   # GitLab API notes 是扁平結構
        })

    # ── 組合回傳 ──────────────────────────────────────────────────────────────
    labels = [
        l["name"] if isinstance(l, dict) else l
        for l in issue_data.get("labels", [])
    ]

    # Assignees（支援多人）
    assignees = [a.get("name", "") for a in issue_data.get("assignees", [])]
    if not assignees and issue_data.get("assignee"):
        assignees = [issue_data["assignee"].get("name", "")]

    # Milestone
    ms = issue_data.get("milestone") or {}
    milestone = {
        "title":      ms.get("title", ""),
        "due_date":   ms.get("due_date", ""),
        "start_date": ms.get("start_date", ""),
    } if ms else {}

    # Time stats（分鐘 → 可讀字串）
    def fmt_duration(secs):
        if not secs:
            return ""
        h, m = divmod(int(secs) // 60, 60)
        return f"{h}h{m:02d}m" if h else f"{m}m"

    ts = issue_data.get("time_stats") or {}
    time_estimate = fmt_duration(ts.get("time_estimate", 0))
    time_spent    = fmt_duration(ts.get("total_time_spent", 0))

    return {
        # ── 識別資訊 ──
        "iid":          issue_data.get("iid", ""),
        "url":          issue_data.get("web_url", ""),
        "title":        issue_data.get("title", "（無標題）"),
        "state":        issue_data.get("state", ""),              # opened / closed
        # ── 人員 ──
        "author":       issue_data.get("author", {}).get("name", "unknown"),
        "assignees":    assignees,                                 # list of names
        # ── 時間 ──
        "created_at":   fmt_time(issue_data.get("created_at", "")),
        "updated_at":   fmt_time(issue_data.get("updated_at", "")),
        "closed_at":    fmt_time(issue_data.get("closed_at", "") or ""),
        "due_date":     issue_data.get("due_date", "") or "",
        # ── 分類 ──
        "labels":       labels,
        "milestone":    milestone,                                 # dict: title, due_date, start_date
        "weight":       issue_data.get("weight", ""),             # None if not used
        # ── 時間追蹤 ──
        "time_estimate": time_estimate,
        "time_spent":    time_spent,
        # ── 內容 ──
        "description":  issue_data.get("description", "") or "",
        "comments":     comments,
    }


# ─── API endpoints ───────────────────────────────────────────────────────

@app.route("/")
def index():
    return send_from_directory(".", "index.html")


# ── 批次爬取（需求 1）：一次處理一個 URL，前端自行排隊呼叫 ──
@app.route("/api/scrape", methods=["POST"])
def api_scrape():
    body     = request.get_json()
    url      = (body.get("url")      or "").strip()
    username = (body.get("username") or "").strip()
    password = (body.get("password") or "").strip()

    if not url or not username or not password:
        return jsonify({"error": "請填寫 URL、帳號、密碼"}), 400

    m = re.match(r"(https?://[^/]+)", url)
    if not m:
        return jsonify({"error": "URL 格式不正確"}), 400
    base_url = m.group(1)

    try:
        issue    = scrape_issue_selenium(base_url, username, password, url)
        raw_text = issue_to_text(issue)

        # ── 需求 2：自動存爬蟲結果 ──
        saved_raw = save_output(raw_text, "raw", url)

        return jsonify({
            "raw_text":  raw_text,
            "issue":     issue,
            "saved_raw": saved_raw,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── GitLab API 版爬取（不需帳密，只需 token）──────────────────────────────
@app.route("/api/scrape_api", methods=["POST"])
def api_scrape_via_api():
    """使用 GitLab REST API 取代 Selenium，速度快 10 倍，不需帳號密碼。"""
    body          = request.get_json()
    url           = (body.get("url")           or "").strip()
    project_id    = body.get("project_id")
    private_token = (body.get("private_token") or "").strip()

    if not url:
        return jsonify({"error": "請填寫 Issue URL"}), 400
    if not project_id:
        return jsonify({"error": "請提供 Project ID（在連線設定填入）"}), 400
    if not private_token:
        return jsonify({"error": "請提供 API Token（在連線設定填入）"}), 400

    m_iid = re.search(r"/issues/(\d+)", url)
    if not m_iid:
        return jsonify({"error": "無法從 URL 解析 Issue 編號"}), 400
    issue_iid = int(m_iid.group(1))

    m_base = re.match(r"(https?://[^/]+)", url)
    if not m_base:
        return jsonify({"error": "URL 格式不正確"}), 400
    base_url = m_base.group(1)

    try:
        issue     = scrape_issue_api(base_url, int(project_id), issue_iid, private_token)
        raw_text  = issue_to_text(issue)
        saved_raw = save_output(raw_text, "raw", url)

        return jsonify({
            "raw_text":  raw_text,
            "issue":     issue,
            "saved_raw": saved_raw,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── LLM 整理（需求 2 + 3）──
@app.route("/api/process", methods=["POST"])
def api_process():
    body          = request.get_json()
    raw_text      = (body.get("raw_text")      or "").strip()
    system_prompt = (body.get("system_prompt") or "").strip()
    url           = (body.get("url")           or "").strip()
    model_name    = (body.get("model_name")    or "").strip()
    model_label   = (body.get("model_label")   or "").strip()
    timeout       = body.get("timeout")
    if timeout is not None:
        try:
            timeout = int(timeout)
        except (ValueError, TypeError):
            timeout = None

    if not raw_text:
        return jsonify({"error": "raw_text 不可為空"}), 400
    if model_name and is_disallowed_model(model_name):
        logger.warning("api_process rejected_flash model=%s url=%s", model_name, url)
        return jsonify({"error": "不接受 Flash 模型"}), 400

    try:
        logger.info("api_process start model=%s model_label=%s timeout=%s raw_text_len=%s url=%s", model_name or "<default>", model_label or "", timeout if timeout is not None else GEMINI_TIMEOUT, len(raw_text), url or "")
        # 需求 3：使用強制結構的 prompt（若使用者有自訂則合併）
        effective_prompt = system_prompt or PROCESS_SYSTEM_PROMPT
        result = call_gemini_cli(
            effective_prompt,
            raw_text,
            timeout=timeout,
            model_name=model_name or None,
        )

        # 需求 3：確保六個區塊都存在
        result = enforce_structure(result)

        # 需求 2：自動存 LLM 整理結果
        saved_result = save_output(result, "result", url, model_name=model_name) if url else None
        logger.info("api_process success model=%s saved_result=%s url=%s", model_name or "<default>", saved_result or "", url or "")

        return jsonify({
            "result": result,
            "saved_result": saved_result,
            "used_model": model_name or None,
            "used_model_label": model_label or None,
        })
    except RuntimeError as e:
        logger.error("api_process runtime_error model=%s url=%s error=%s", model_name or "<default>", url or "", str(e))
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        logger.exception("api_process unexpected_error model=%s url=%s", model_name or "<default>", url or "")
        return jsonify({"error": f"未知錯誤：{e}"}), 500


@app.route("/api/export", methods=["POST"])
def api_export():
    body           = request.get_json()
    processed_text = (body.get("processed_text") or "").strip()
    export_prompt  = (body.get("export_prompt")  or "").strip()

    if not processed_text:
        return jsonify({"error": "processed_text 不可為空"}), 400

    default_export_prompt = (
        "請將以上整理好的 Issue 摘要，轉換成以下 JSON 格式輸出（只輸出 JSON，不加說明）：\n"
        '{\n'
        '  "title": "Issue 標題",\n'
        '  "summary": "一句話摘要",\n'
        '  "root_cause": "根本原因",\n'
        '  "solution": "解決方式",\n'
        '  "action_items": ["待辦1", "待辦2"],\n'
        '  "sentiment": "positive|neutral|negative|mixed"\n'
        '}'
    )

    export_system = "你是資料格式化助理，請嚴格按照使用者指定的格式輸出，不加任何多餘說明。"
    full_user = f"{processed_text}\n\n---\n\n{export_prompt or default_export_prompt}"

    try:
        output = call_gemini_cli(export_system, full_user)
        return jsonify({"output": output})
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        return jsonify({"error": f"未知錯誤：{e}"}), 500


# ── 列出歷史存檔（需求 2）──
@app.route("/api/outputs", methods=["GET"])
def api_outputs():
    return jsonify({"files": list_outputs()})


# ── 讀取單一存檔內容 ──
@app.route("/api/outputs/<filename>", methods=["GET"])
def api_output_file(filename):
        # Security: only allow .txt and .xlsx files
    if not re.match(r"^[\w\-\.]+\.(txt|xlsx)$", filename):
        return jsonify({"error": "非法檔名"}), 400
    # Search in raw/, results/, and excel/ directories
    for subdir in [OUTPUT_RAW, OUTPUT_RESULTS, OUTPUT_EXCEL]:
        filepath = subdir / filename
        if filepath.exists():
            if filename.endswith(".xlsx"):
                return send_from_directory(
                    str(subdir.resolve()), filename, as_attachment=True
                )
            return jsonify({"filename": filename, "content": filepath.read_text(encoding="utf-8")})
    return jsonify({"error": "檔案不存在"}), 404


# ── Label 解析（對應 gitlab_to_excel.py 的業務邏輯）────────────────────────
PRIORITY_MAP = {
    "Priority::High":   "High",
    "Priority::Medium": "Medium",
    "Priority::Low":    "Low",
}
TAG_KEYWORDS   = ["PES", "Suggestion", "Bug"]
TEAM_MAPPING   = {
    "Team::UI/UX Design":   "UI/UX",
    "Team::Frontend":       "FE",
    "Team::Backend":        "BE",
    "Team::Infra":          "Infra",
    "Team::AI/SAM worker":  "AI worker",
    "Team::AI":             "AI",
    "Team::IT":             "IE",
}
IGNORED_LABELS = {"Enhancement", "UI Done", "UX Done"}

def parse_labels(labels: list) -> dict:
    """
    與 gitlab_to_excel.py 相同的 label 解析邏輯，
    輸入 label 名稱清單，回傳結構化 dict。
    """
    label_set = set(labels)

    # Priority
    priority = ""
    for full, short in PRIORITY_MAP.items():
        if full in label_set:
            priority = short
            break

    # Tag
    lower_tags = {t.lower() for t in TAG_KEYWORDS}
    found_tags = [kw for kw in TAG_KEYWORDS
                  if any(kw.lower() == l.lower() for l in labels)]

    # Epics
    epics = [l.replace("Epics:", "").strip()
             for l in labels if l.startswith("Epics:")]

    # Team columns（UI/UX 特規：需同時有 UI Done + UX Done 才算 ✓）
    known_team = set(TEAM_MAPPING.keys())
    team_status = {col: "" for col in TEAM_MAPPING.values()}
    for label, col in TEAM_MAPPING.items():
        if label in label_set:
            if col == "UI/UX":
                team_status[col] = "✓" if ("UI Done" in label_set and "UX Done" in label_set) else "0%"
            else:
                team_status[col] = "0%"

    # 其他標籤
    known_priority = set(PRIORITY_MAP.keys())
    other = [l for l in labels
             if l not in known_priority
             and l not in known_team
             and l not in IGNORED_LABELS
             and not l.startswith("Epics:")
             and l.lower() not in lower_tags]

    return {
        "priority":   priority,
        "tags":       ", ".join(found_tags),
        "epics":      ", ".join(epics),
        "other_tags": ", ".join(other),
        **team_status,   # UI/UX, FE, BE, Infra, AI worker, AI, IE
    }


def issue_to_excel_row(issue: dict) -> dict:
    """
    把 scrape_issue_api() 回傳的完整 issue dict
    轉換為 Excel 一列用的扁平 dict。
    """
    parsed = parse_labels(issue.get("labels", []))
    ms = issue.get("milestone") or {}
    return {
        "Issue ID":     issue.get("iid", ""),
        "Title":        issue.get("title", ""),
        "State":        issue.get("state", ""),
        "Priority":     parsed["priority"],
        "Tag":          parsed["tags"],
        "Epics":        parsed["epics"],
        "其他標籤":     parsed["other_tags"],
        "UI/UX":        parsed.get("UI/UX", ""),
        "FE":           parsed.get("FE", ""),
        "BE":           parsed.get("BE", ""),
        "Infra":        parsed.get("Infra", ""),
        "AI worker":    parsed.get("AI worker", ""),
        "AI":           parsed.get("AI", ""),
        "IE":           parsed.get("IE", ""),
        "指派對象":     ", ".join(issue.get("assignees", [])),
        "建立者":       issue.get("author", ""),
        "建立時間":     issue.get("created_at", ""),
        "最後更新":     issue.get("updated_at", ""),
        "Due Date":     issue.get("due_date", ""),
        "預計完成版本": ms.get("title", ""),
        "Weight":       "" if issue.get("weight") is None else issue.get("weight", ""),
        "預估工時":     issue.get("time_estimate", ""),
        "實花工時":     issue.get("time_spent", ""),
        "URL":          issue.get("url", ""),
    }


# ── GitLab 篩選頁面 URL → Issue URL 清單 ──────────────────────────────
def _parse_filter_url(filter_url: str):
    """
    解析 GitLab 議題篩選頁面 URL。
    回傳 (base_url, project_path, api_params, labels, or_labels)
    """
    from urllib.parse import urlparse, parse_qs, unquote

    parsed   = urlparse(filter_url)
    base_url = f"{parsed.scheme}://{parsed.netloc}"

    # 取出 project path
    path = parsed.path
    m = re.match(r"^(/[^/].+?)/-/issues", path)
    if not m:
        raise ValueError(f"無法從 URL 解析 project path：{path}")
    project_path = m.group(1).lstrip("/")

    # parse_qs 解析（keep_blank_values=False）
    # 注意：瀏覽器傳來的 query string 已經是 percent-encoded，
    # parse_qs 預設會幫我們 decode，所以 key 會是 or[label_name][]
    qs = parse_qs(parsed.query, keep_blank_values=False)

    api_params = {}

    # state
    api_params["state"] = qs.get("state", ["opened"])[0]

    # milestone — GitLab REST API 用 "milestone"（網頁 URL 用 milestone_title，但 API 參數名不同）
    if "milestone_title" in qs:
        api_params["milestone"] = qs["milestone_title"][0]

    # assignee
    if "assignee_username" in qs:
        api_params["assignee_username"] = qs["assignee_username"][0]

    # AND labels（label_name[]）
    labels = qs.get("label_name[]", [])

    # OR labels：key 可能是 or[label_name][] 或未解碼的 or%5Blabel_name%5D%5B%5D
    or_labels = qs.get("or[label_name][]", [])

    # 若以上取不到，嘗試逐 key 搜尋（防禦性解碼）
    if not or_labels:
        for k, v in qs.items():
            if "label_name" in unquote(k) and "or" in unquote(k).lower():
                or_labels = v
                break

    # NOT labels（not[label_name][]）— 排除含特定 label 的 issue
    not_labels = qs.get("not[label_name][]", [])
    if not not_labels:
        for k, v in qs.items():
            if "label_name" in unquote(k) and "not" in unquote(k).lower():
                not_labels = v
                break

    return base_url, project_path, api_params, labels, or_labels, not_labels


def _fetch_issues_from_api(
    base_url: str,
    project_path: str,
    api_params: dict,
    labels: list,
    or_labels: list,
    not_labels: list = None,
    project_id: int = None,
    private_token: str = None,
) -> list[dict]:
    """
    呼叫 GitLab API 取得 Issue 清單。
    使用 requests 套件（比 urllib 更穩定，支援 http://IP:port）。
    OR label 邏輯：分批查詢各 label 後合併去重。
    自動翻頁。支援 Private Access Token 認證。
    """
    try:
        import requests as req_lib
    except ImportError:
        raise RuntimeError(
            "缺少 requests 套件，請執行：pip install requests"
        )

    # 直接使用傳入的 project_id（數字），不需要搜尋
    api_base = f"{base_url}/api/v4/projects/{project_id}/issues"

    # 若有提供 token，加入認證 header
    headers = {}
    if private_token:
        headers["PRIVATE-TOKEN"] = private_token

    def fetch_all_pages(params: dict) -> list[dict]:
        results = []
        page = 1
        while True:
            p = {**params, "page": page, "per_page": 100}
            try:
                resp = gitlab_api_get(req_lib, api_base, params=p, headers=headers, timeout=15)
            except Exception as e:
                raise RuntimeError(f"GitLab API 連線失敗：{e}")

            if resp.status_code == 404:
                raise RuntimeError(
                    f"404 找不到專案（project_id={project_id}），"
                    f"請確認 Project ID 正確。實際呼叫 URL：{resp.url}"
                )
            if resp.status_code == 401:
                raise RuntimeError(
                    "401 未授權：此 GitLab 需要 Personal Access Token，"
                    "請在「快速匯入」區塊填入 API Token"
                )
            if not resp.ok:
                raise RuntimeError(
                    f"API 回傳錯誤 {resp.status_code}：{resp.text[:200]}"
                )

            data = resp.json()
            if not data:
                break
            results.extend(data)
            if len(data) < 100:
                break
            page += 1
        return results

    all_issues = {}

    # 若有 NOT labels，加入 not[labels] 排除條件（GitLab API 支援）
    not_params = {}
    if not_labels:
        not_params["not[labels]"] = ",".join(not_labels)

    if or_labels:
        for label in or_labels:
            params = {**api_params, **not_params}
            if labels:
                params["labels"] = label + "," + ",".join(labels)
            else:
                params["labels"] = label
            for issue in fetch_all_pages(params):
                all_issues[issue["iid"]] = issue
    else:
        params = {**api_params, **not_params}
        if labels:
            params["labels"] = ",".join(labels)
        for issue in fetch_all_pages(params):
            all_issues[issue["iid"]] = issue

    return sorted(all_issues.values(), key=lambda x: x["iid"])


def _format_issue_preview(issue: dict) -> dict:
    """將 GitLab API 回傳的 issue 物件格式化為預覽用的精簡結構。"""
    assignees = [
        a.get("name") or a.get("username", "")
        for a in issue.get("assignees", [])
    ]
    # 相容 assignee（單一）與 assignees（陣列）兩種格式
    if not assignees and issue.get("assignee"):
        assignees = [issue["assignee"].get("name") or issue["assignee"].get("username", "")]

    milestone = None
    if issue.get("milestone"):
        milestone = {
            "title":    issue["milestone"].get("title", ""),
            "due_date": issue["milestone"].get("due_date", ""),
        }

    return {
        "iid":       issue["iid"],
        "title":     issue.get("title", ""),
        "web_url":   issue.get("web_url", ""),
        "state":     issue.get("state", ""),
        "assignees": assignees,
        "milestone": milestone,
        "labels":    [l["name"] if isinstance(l, dict) else l
                      for l in issue.get("labels", [])],
    }


@app.route("/api/preview_issues", methods=["POST"])
def api_preview_issues():
    """
    接收多個 Issue URL，輕量呼叫 GitLab API 取得基本預覽資訊。
    （iid / title / state / assignees / milestone）
    不取留言，速度快。
    """
    try:
        import requests as req_lib
    except ImportError:
        return jsonify({"error": "缺少 requests 套件，請執行：pip install requests"}), 500

    body          = request.get_json(force=True)
    urls          = [u.strip() for u in (body.get("urls") or []) if u.strip()]
    project_id    = body.get("project_id")
    private_token = (body.get("private_token") or "").strip() or None

    if not urls:
        return jsonify({"error": "請提供至少一個 Issue URL"}), 400

    headers = {"PRIVATE-TOKEN": private_token} if private_token else {}

    # 從第一個 URL 推斷 base_url
    m_base = re.match(r"(https?://[^/]+)", urls[0])
    if not m_base:
        return jsonify({"error": "無法解析 base URL"}), 400
    base_url = m_base.group(1)

    if not project_id:
        return jsonify({"error": "請先在連線設定填寫 Project ID"}), 400

    api_base = f"{base_url}/api/v4/projects/{project_id}/issues"
    issues   = []
    errors   = []

    for url in urls:
        m_iid = re.search(r"/issues/(\d+)", url)
        if not m_iid:
            errors.append({"url": url, "error": "無法解析 issue IID"})
            continue
        iid = int(m_iid.group(1))
        try:
            resp = gitlab_api_get(req_lib, f"{api_base}/{iid}", headers=headers, timeout=10)
            if resp.status_code == 401:
                return jsonify({"error": "401 未授權，請確認 API Token 有效"}), 401
            if resp.status_code == 404:
                errors.append({"url": url, "error": f"找不到 Issue #{iid}"})
                continue
            if not resp.ok:
                errors.append({"url": url, "error": f"API 錯誤 {resp.status_code}"})
                continue
            issues.append(_format_issue_preview(resp.json()))
        except Exception as e:
            errors.append({"url": url, "error": str(e)})

    return jsonify({
        "count":  len(issues),
        "issues": issues,
        "errors": errors,
    })


@app.route("/api/resolve_filter_url", methods=["POST"])
def api_resolve_filter_url():
    """
    接收 GitLab 議題篩選頁面 URL，
    回傳符合條件的所有 Issue web_url 清單。
    """
    body          = request.get_json()
    filter_url    = (body.get("filter_url")    or "").strip()
    project_id    = body.get("project_id")  # 選填的數字 ID，若提供則直接使用
    private_token = (body.get("private_token") or "").strip()  # 選填的 PAT

    if not filter_url:
        return jsonify({"error": "請提供篩選頁面 URL"}), 400

    try:
        base_url, project_path, api_params, labels, or_labels, not_labels = _parse_filter_url(filter_url)
        issues = _fetch_issues_from_api(
            base_url, project_path, api_params, labels, or_labels,
            not_labels=not_labels,
            project_id=project_id,
            private_token=private_token or None,
        )

        issue_list = [_format_issue_preview(issue) for issue in issues]

        return jsonify({
            "count":    len(issue_list),
            "project":  project_path,
            "issues":   issue_list,
            "web_urls": [i["web_url"] for i in issue_list],
        })

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        return jsonify({"error": f"未知錯誤：{e}"}), 500


# ── 批次匯出 Excel ────────────────────────────────────────────────────────
@app.route("/api/batch_export_excel", methods=["POST"])
def api_batch_export_excel():
    """
    接收 issue URL 清單 + project_id + private_token，
    逐一呼叫 scrape_issue_api() 取得完整資料，
    套用 issue_to_excel_row() 解析 label，
    產出格式化 Excel (.xlsx) 並回傳檔名。
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "缺少 openpyxl，請執行：pip install openpyxl"}), 500

    body          = request.get_json()
    urls          = [u.strip() for u in (body.get("urls") or []) if u.strip()]
    project_id    = body.get("project_id")
    private_token = (body.get("private_token") or "").strip() or None

    if not urls:
        return jsonify({"error": "請提供至少一個 Issue URL"}), 400
    if not project_id:
        return jsonify({"error": "請提供 Project ID"}), 400

    # 從第一個 URL 推斷 base_url
    m_base = re.match(r"(https?://[^/]+)", urls[0])
    if not m_base:
        return jsonify({"error": "無法解析 base URL"}), 400
    base_url = m_base.group(1)

    rows   = []
    errors = []

    for url in urls:
        m_iid = re.search(r"/issues/(\d+)", url)
        if not m_iid:
            errors.append({"url": url, "error": "無法解析 issue IID"})
            continue
        issue_iid = int(m_iid.group(1))
        try:
            issue = scrape_issue_api(base_url, int(project_id), issue_iid, private_token)
            rows.append(issue_to_excel_row(issue))
        except Exception as e:
            errors.append({"url": url, "error": str(e)})

    if not rows:
        return jsonify({"error": "所有 issue 均抓取失敗", "details": errors}), 500

    # ── 建立 Excel ───────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Issues"

    COLS = [
        "Issue ID", "Title", "State", "Priority", "Tag", "Epics", "其他標籤",
        "UI/UX", "FE", "BE", "Infra", "AI worker", "AI", "IE",
        "指派對象", "建立者", "建立時間", "最後更新",
        "Due Date", "預計完成版本", "Weight", "預估工時", "實花工時", "URL",
    ]

    # 欄寬設定
    col_widths = {
        "Issue ID": 10, "Title": 45, "State": 10, "Priority": 10,
        "Tag": 14, "Epics": 20, "其他標籤": 22,
        "UI/UX": 8, "FE": 8, "BE": 8, "Infra": 8, "AI worker": 10, "AI": 8, "IE": 8,
        "指派對象": 16, "建立者": 14, "建立時間": 16, "最後更新": 16,
        "Due Date": 13, "預計完成版本": 16, "Weight": 8,
        "預估工時": 10, "實花工時": 10, "URL": 60,
    }

    # 標頭樣式
    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill  = PatternFill("solid", start_color="2F4F8F")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for ci, col_name in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font    = header_fill_font = header_font
        cell.fill    = header_fill
        cell.alignment = header_align
        cell.border  = thin_border
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col_name, 14)

    ws.row_dimensions[1].height = 30

    # 資料列樣式（交錯底色）
    data_font  = Font(name="Arial", size=10)
    fill_even  = PatternFill("solid", start_color="EEF2F8")
    fill_odd   = PatternFill("solid", start_color="FFFFFF")
    data_align = Alignment(vertical="center", wrap_text=False)
    url_font   = Font(name="Arial", size=10, color="1155CC", underline="single")

    priority_colors = {"High": "FF4444", "Medium": "FF9900", "Low": "44AA44"}

    for ri, row_data in enumerate(rows, start=2):
        fill = fill_even if ri % 2 == 0 else fill_odd
        for ci, col_name in enumerate(COLS, start=1):
            value = row_data.get(col_name, "")
            cell  = ws.cell(row=ri, column=ci, value=value)
            cell.border    = thin_border
            cell.alignment = data_align
            cell.fill      = fill

            # URL 欄 → 藍色
            if col_name == "URL" and value:
                cell.font = url_font
            # Priority 欄 → 彩色
            elif col_name == "Priority" and value in priority_colors:
                cell.font = Font(name="Arial", size=10, bold=True,
                                 color=priority_colors[value])
            else:
                cell.font = data_font

    # 凍結首列
    ws.freeze_panes = "A2"

    # 自動篩選
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    # 儲存至 outputs/excel/
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"excel_{ts}_issues.xlsx"
    filepath = OUTPUT_EXCEL / filename
    wb.save(str(filepath))

    return jsonify({
        "filename": filename,
        "count":    len(rows),
        "errors":   errors,
    })


# ── Prompt 模板管理 ──────────────────────────────────────────────────────

@app.route("/api/prompts", methods=["GET"])
def api_list_prompts():
    """列出 prompts/ 資料夾內所有 .md 檔"""
    prompts = []
    for f in sorted(PROMPTS_DIR.glob("*.md"), key=lambda x: x.name):
        prompts.append({
            "filename": f.name,
            "name":     f.stem,
            "mtime":    datetime.fromtimestamp(f.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
            "size":     f.stat().st_size,
        })
    return jsonify({"prompts": prompts})


@app.route("/api/prompts/<filename>", methods=["GET"])
def api_get_prompt(filename):
    """讀取單一 prompt 檔案內容"""
    filepath = (PROMPTS_DIR / filename).resolve()
    if not str(filepath).startswith(str(PROMPTS_DIR.resolve())):
        return jsonify({"error": "非法的檔案路徑"}), 400
    if not filepath.exists():
        return jsonify({"error": f"找不到 prompt：{filename}"}), 404
    content = filepath.read_text(encoding="utf-8")
    return jsonify({"filename": filename, "name": filepath.stem, "content": content})


@app.route("/api/prompts", methods=["POST"])
def api_save_prompt():
    """建立或覆蓋 prompt 檔案"""
    data      = request.get_json(force=True)
    filename  = (data.get("filename") or "").strip()
    content   = (data.get("content")  or "").strip()
    overwrite = bool(data.get("overwrite", False))

    if not filename:
        return jsonify({"error": "filename 不可為空"}), 400
    if not content:
        return jsonify({"error": "content 不可為空"}), 400
    if not filename.endswith(".md"):
        filename += ".md"

    # 防止路徑穿越攻擊
    filepath = (PROMPTS_DIR / filename).resolve()
    if not str(filepath).startswith(str(PROMPTS_DIR.resolve())):
        return jsonify({"error": "非法的檔案路徑"}), 400

    if filepath.exists() and not overwrite:
        return jsonify({"error": f"檔案已存在：{filename}，請使用覆蓋儲存"}), 409

    filepath.write_text(content, encoding="utf-8")
    return jsonify({"filename": filename, "saved": True})


@app.route("/api/prompts/<filename>", methods=["DELETE"])
def api_delete_prompt(filename):
    """刪除 prompt 檔案"""
    filepath = (PROMPTS_DIR / filename).resolve()
    if not str(filepath).startswith(str(PROMPTS_DIR.resolve())):
        return jsonify({"error": "非法的檔案路徑"}), 400
    if not filepath.exists():
        return jsonify({"error": f"找不到 prompt：{filename}"}), 404
    filepath.unlink()
    return jsonify({"filename": filename, "deleted": True})


# ── 健康檢查 ──
@app.route("/api/health", methods=["GET"])
def api_health():
    cli_path  = _resolve_gemini_executable()
    available = os.path.isfile(cli_path) or bool(shutil.which(GEMINI_CLI))

    return jsonify({
        "gemini_cli":      GEMINI_CLI,
        "cli_found":       available,
        "cli_path":        str(cli_path),
        "timeout":         GEMINI_TIMEOUT,
        "max_input_chars":    MAX_INPUT_CHARS,
        "output_dir":         str(OUTPUT_DIR.resolve()),
        "output_dir_raw":     str(OUTPUT_RAW.resolve()),
        "output_dir_results": str(OUTPUT_RESULTS.resolve()),
        "output_dir_excel":   str(OUTPUT_EXCEL.resolve()),
        "model_chain":        get_model_chain(),
    })


@app.route("/api/probe_models", methods=["POST"])
def api_probe_models():
    body = request.get_json(force=True) or {}
    models = body.get("models") or []
    timeout = body.get("timeout")

    if not isinstance(models, list) or not models:
        return jsonify({"error": "models 必須為非空陣列"}), 400

    cleaned_models = []
    for model in models:
        name = str(model).strip()
        if name:
            cleaned_models.append(name)

    if not cleaned_models:
        return jsonify({"error": "models 必須包含至少一個有效模型名稱"}), 400

    logger.info("api_probe_models start models=%s timeout=%s", cleaned_models, timeout if timeout else GEMINI_PROBE_TIMEOUT)
    results = [probe_gemini_model(model_name, timeout=timeout) for model_name in cleaned_models]
    logger.info("api_probe_models done results=%s", [{k: r.get(k) for k in ("model", "ok", "status", "returncode")} for r in results])
    return jsonify({
        "cli_path": str(_resolve_gemini_executable()),
        "probe_timeout": timeout if (timeout and timeout > 0) else GEMINI_PROBE_TIMEOUT,
        "results": results,
    })


if __name__ == "__main__":
    debug_mode = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(host=FLASK_HOST, port=FLASK_PORT, debug=debug_mode, use_reloader=debug_mode)
